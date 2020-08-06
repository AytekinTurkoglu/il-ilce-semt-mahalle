import pymysql.cursors
from openpyxl import load_workbook
import match

connection = pymysql.connect(host='localhost',
                             user='root',
                             password='',
                             db='deneme',
                             charset='utf8',
                             cursorclass=pymysql.cursors.DictCursor)
cursor = connection.cursor()

try:
    wb = load_workbook('data/pk_list_05.08.2020.xlsx')
    ws = wb.active

    lower_map = {ord(u'I'): u'ı'}
    regions = {}
    for index, row in enumerate(ws.rows):
        if index == 0:
            continue

        index_str = str(index + 1)

        city = ws["A" + index_str].value
        district = ws["B" + index_str].value
        neighborhood = ws["D" + index_str].value
        postal_code = ws["E" + index_str].value

        if city not in regions:
            regions[city] = {}

        if district not in regions[city]:
            regions[city][district] = {}

        if neighborhood not in regions[city][district]:
            regions[city][district][neighborhood] = {}

        regions[city][district][neighborhood] = postal_code

    for city in regions.keys():
        city_name = city.translate(lower_map).title().strip().replace("i̇", "i")

        with connection.cursor() as cursor:
            city_id = match.city[city_name]

            sql = "INSERT INTO `City` (`Id`, `Name`) VALUES (%s, %s);"
            cursor.execute(sql, (city_id, city_name))
            connection.commit()

            for district in regions[city].keys():
                district_name = district.translate(lower_map).title().strip().replace("i̇", "i")

                with connection.cursor() as cursor:
                    sql = "INSERT INTO `District` (`CityId`, `Name`) VALUES (%s, %s);"
                    cursor.execute(sql, (city_id, district_name))
                    connection.commit()
                    district_id = cursor.lastrowid

                    for neighborhood in regions[city][district].keys():
                        neighborhood_name = neighborhood.translate(lower_map).title().strip().replace("i̇", "i")

                        with connection.cursor() as cursor:
                            sql = "INSERT INTO `Neighborhood` (`DistrictId`, `Name`, `PostalCode`) VALUES (%s, %s, %s);"
                            cursor.execute(sql, (
                                district_id, neighborhood_name, regions[city][district][neighborhood]))
                            connection.commit()
                            part_id = cursor.lastrowid
                            postal_code = regions[city][district][neighborhood]

                            print(
                                city_name + ' / ' + district_name + ' / ' + neighborhood_name + ' / ' + str(
                                    postal_code))

finally:
    connection.close()
